import openpyxl
from django.http import HttpResponse
from apps.admin_inventory.models import Booklist, Inventory, InventoryHistory

def export_listOfAcquisitions(request):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Report"

    headers = [
        "itemcallnumber", "ccode", "itype", "title", "subtitle", 
        "author", "copyrightdate", "publishercode", "barcode", 
        "isbn", "dateaccessioned", "copynumber", "volume", 
        "editionstatement", "paidfor", "price", "booksellerid",
        "bookstate"
    ]
    sheet.append(headers)

    books = Booklist.objects.all()

    for book in books:
        book_statuses = Inventory.objects.filter(book=book)
        book_status = book_statuses.first().get_status_display()
        
        sheet.append([
            book.item_call_num,   # itemcallnumber
            book.col_code,        # ccode
            book.itype,           # itype
            book.title,           # title
            book.subtitle,        # subtitle
            book.author,          # author
            book.copyrightdate,   # copyrightdate
            book.publisher_code,  # publishercode
            book.barcode,         # barcode
            book.isbn,            # isbn
            book.date_accessioned, # dateaccessioned
            book.copy_num,        # copynumber
            book.volume,          # volume
            book.edition_stmt,    # editionstatement
            book.paidfor,         # paidfor
            book.price,           # price
            book.bookseller_id,   # booksellerid
            book_status,          # book status
        ])

    # Response headers for downloading the file
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = "attachment; filename=list-of-acquisitions.xlsx"

    workbook.save(response)
    return response

def export_noBarcodeTag(request):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Report"

    headers = [
        "itemcallnumber", "ccode", "itype", "title", "subtitle", 
        "author", "copyrightdate", "publishercode", "barcode", 
        "isbn", "dateaccessioned", "copynumber", "volume", 
        "editionstatement", "paidfor", "price", "booksellerid", "remarks"
    ]
    sheet.append(headers)

    books = Booklist.objects.filter(inventories__status=Inventory.NO_BARCODE_TAG)

    for book in books:
        inventory = book.inventories.filter(status=Inventory.NO_BARCODE_TAG).first()

        remark = None
        if inventory:
            history = inventory.history.order_by('-datetime_checked').first()
            remark = history.remarks if history else None
        
        # Append data to the sheet
        sheet.append([
            book.item_call_num,   # itemcallnumber
            book.col_code,        # ccode
            book.itype,           # itype
            book.title,           # title
            book.subtitle,        # subtitle
            book.author,          # author
            book.copyrightdate,   # copyrightdate
            book.publisher_code,  # publishercode
            book.barcode,         # barcode
            book.isbn,            # isbn
            book.date_accessioned, # dateaccessioned
            book.copy_num,        # copynumber
            book.volume,          # volume
            book.edition_stmt,    # editionstatement
            book.paidfor,         # paidfor
            book.price,           # price
            book.bookseller_id,   # booksellerid
            remark,               # remarks
        ])

    # Response headers for downloading the file
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = "attachment; filename=no-barcode-tag.xlsx"

    workbook.save(response)
    return response


def export_forRepair(request):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Report"

    headers = [
        "itemcallnumber", "ccode", "itype", "title", "subtitle", 
        "author", "copyrightdate", "publishercode", "barcode", 
        "isbn", "dateaccessioned", "copynumber", "volume", 
        "editionstatement", "paidfor", "price", "booksellerid", "remarks"
    ]
    sheet.append(headers)

    books = Booklist.objects.filter(inventories__status=Inventory.FOR_REPAIR)

    for book in books:
        inventory = book.inventories.filter(status=Inventory.FOR_REPAIR).first()
        remark = None
        if inventory:
            history = inventory.history.order_by('-datetime_checked').first()
            remark = history.remarks if history else None

        sheet.append([
            book.item_call_num,   # itemcallnumber
            book.col_code,        # ccode
            book.itype,           # itype
            book.title,           # title
            book.subtitle,        # subtitle
            book.author,          # author
            book.copyrightdate,   # copyrightdate
            book.publisher_code,  # publishercode
            book.barcode,         # barcode
            book.isbn,            # isbn
            book.date_accessioned, # dateaccessioned
            book.copy_num,        # copynumber
            book.volume,          # volume
            book.edition_stmt,    # editionstatement
            book.paidfor,         # paidfor
            book.price,           # price
            book.bookseller_id,   # booksellerid
            remark,               # remarks
        ])

    # Response headers for downloading the file
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = "attachment; filename=for-repair.xlsx"

    workbook.save(response)
    return response

def export_forDisposal(request):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Report"

    headers = [
        "itemcallnumber", "ccode", "itype", "title", "subtitle", 
        "author", "copyrightdate", "publishercode", "barcode", 
        "isbn", "dateaccessioned", "copynumber", "volume", 
        "editionstatement", "paidfor", "price", "booksellerid", "remarks"
    ]
    sheet.append(headers)

    books = Booklist.objects.filter(inventories__status=Inventory.FOR_DISPOSAL)

    for book in books:
        inventory = book.inventories.filter(status=Inventory.FOR_DISPOSAL).first()
        remark = None
        if inventory:
            history = inventory.history.order_by('-datetime_checked').first()
            remark = history.remarks if history else None

        sheet.append([
            book.item_call_num,   # itemcallnumber
            book.col_code,        # ccode
            book.itype,           # itype
            book.title,           # title
            book.subtitle,        # subtitle
            book.author,          # author
            book.copyrightdate,   # copyrightdate
            book.publisher_code,  # publishercode
            book.barcode,         # barcode
            book.isbn,            # isbn
            book.date_accessioned, # dateaccessioned
            book.copy_num,        # copynumber
            book.volume,          # volume
            book.edition_stmt,    # editionstatement
            book.paidfor,         # paidfor
            book.price,           # price
            book.bookseller_id,   # booksellerid
            remark,               # remarks
        ])

    # Response headers for downloading the file
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = "attachment; filename=for-disposal.xlsx"

    workbook.save(response)
    return response

def export_notFound(request):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Report"

    headers = [
        "itemcallnumber", "ccode", "itype", "title", "subtitle", 
        "author", "copyrightdate", "publishercode", "barcode", 
        "isbn", "dateaccessioned", "copynumber", "volume", 
        "editionstatement", "paidfor", "price", "booksellerid",
    ]
    sheet.append(headers)

    books = Booklist.objects.filter(inventories__status=Inventory.NOT_FOUND)

    for book in books:
        sheet.append([
            book.item_call_num,   # itemcallnumber
            book.col_code,        # ccode
            book.itype,           # itype
            book.title,           # title
            book.subtitle,        # subtitle
            book.author,          # author
            book.copyrightdate,   # copyrightdate
            book.publisher_code,  # publishercode
            book.barcode,         # barcode
            book.isbn,            # isbn
            book.date_accessioned, # dateaccessioned
            book.copy_num,        # copynumber
            book.volume,          # volume
            book.edition_stmt,    # editionstatement
            book.paidfor,         # paidfor
            book.price,           # price
            book.bookseller_id,   # booksellerid
        ])

    # Response headers for downloading the file
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = "attachment; filename=not-found.xlsx"

    workbook.save(response)
    return response

def export_Found(request):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Report"

    headers = [
        "itemcallnumber", "ccode", "itype", "title", "subtitle", 
        "author", "copyrightdate", "publishercode", "barcode", 
        "isbn", "dateaccessioned", "copynumber", "volume", 
        "editionstatement", "paidfor", "price", "booksellerid",
    ]
    sheet.append(headers)

    books = Booklist.objects.filter(inventories__status__in=[
        Inventory.GOOD_CONDITION, 
        Inventory.NO_BARCODE_TAG, 
        Inventory.FOR_REPAIR, 
        Inventory.FOR_DISPOSAL
    ])

    for book in books:
        sheet.append([
            book.item_call_num,   # itemcallnumber
            book.col_code,        # ccode
            book.itype,           # itype
            book.title,           # title
            book.subtitle,        # subtitle
            book.author,          # author
            book.copyrightdate,   # copyrightdate
            book.publisher_code,  # publishercode
            book.barcode,         # barcode
            book.isbn,            # isbn
            book.date_accessioned, # dateaccessioned
            book.copy_num,        # copynumber
            book.volume,          # volume
            book.edition_stmt,    # editionstatement
            book.paidfor,         # paidfor
            book.price,           # price
            book.bookseller_id,   # booksellerid
        ])

    # Response headers for downloading the file
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = "attachment; filename=found.xlsx"

    workbook.save(response)
    return response
